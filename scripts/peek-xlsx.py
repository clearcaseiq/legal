import openpyxl

files = [
    r"C:\Business\ClearCaseIQ Inc\Legal\Attoney Data\all_attys_2026-7.xlsx",
    r"C:\Business\ClearCaseIQ Inc\Legal\Attoney Data\atty_practicearea_2026-7.xlsx",
    r"C:\Business\ClearCaseIQ Inc\Legal\Attoney Data\atty_specialties_2026-7.xlsx",
    r"C:\Business\ClearCaseIQ Inc\Legal\Attoney Data\cla_sections_2026-7.xlsx",
    r"C:\Business\ClearCaseIQ Inc\Legal\Attoney Data\discipline_2026-7.xlsx",
]
for f in files:
    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_row=3, values_only=True))
    name = f.split("\\")[-1]
    print(f"=== {name} ===")
    if rows:
        print(f"  Columns: {rows[0]}")
    if len(rows) > 1:
        print(f"  Row 1:   {rows[1]}")
    if len(rows) > 2:
        print(f"  Row 2:   {rows[2]}")
    print(f"  Total rows: {ws.max_row - 1}")
    print()
    wb.close()
