"""
Convert CPRA .xlsx files to .csv one at a time using openpyxl read-only mode.

Usage:
  python scripts/xlsx-to-csv.py <xlsx_file>
  
Example:
  python scripts/xlsx-to-csv.py "C:\path\to\all_attys_2026-7.xlsx"
"""
import sys, os, csv, gc

os.environ['OPENBLAS_NUM_THREADS'] = '1'
os.environ['OMP_NUM_THREADS'] = '1'

from openpyxl import load_workbook

def convert(xlsx_path):
    csv_path = xlsx_path.rsplit('.', 1)[0] + '.csv'
    name = os.path.basename(xlsx_path)
    print(f"Converting {name}...", flush=True)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        count = 0
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(c).strip() if c is not None else '' for c in row])
            count += 1
            if count % 50000 == 0:
                print(f"  {count:,} rows...", flush=True)
    
    wb.close()
    gc.collect()
    print(f"  Done: {count:,} rows -> {os.path.basename(csv_path)}")
    return csv_path

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python xlsx-to-csv.py <file.xlsx>")
        sys.exit(1)
    convert(sys.argv[1])
