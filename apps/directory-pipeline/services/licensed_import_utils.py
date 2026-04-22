"""Shared helpers for compliant licensed/manual dataset imports."""

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_ALIASES = {
    "row_id": ["row_id", "record_id", "id", "attorney_id"],
    "full_name": ["full_name", "name", "attorney_name", "display_name"],
    "first_name": ["first_name", "firstname", "given_name"],
    "last_name": ["last_name", "lastname", "surname", "family_name"],
    "bar_number": ["bar_number", "registration_number", "attorney_registration_number", "license_number"],
    "license_status": ["license_status", "status", "registration_status"],
    "admission_date": ["admission_date", "admitted_at", "date_of_admission"],
    "firm_name": ["firm_name", "company", "employer", "organization", "law_firm"],
    "phone": ["phone", "phone_number", "office_phone"],
    "email": ["email", "email_address"],
    "website": ["website", "website_url", "firm_website"],
    "address_1": ["address_1", "address1", "street", "street_1", "line1"],
    "address_2": ["address_2", "address2", "street_2", "line2"],
    "city": ["city", "town"],
    "state": ["state", "state_code"],
    "zip": ["zip", "zip_code", "postal_code"],
    "practice_areas": ["practice_areas", "practice_area", "specialties", "specializations"],
    "certifications": ["certifications", "board_certifications"],
    "bio_summary": ["bio_summary", "bio", "summary", "notes"],
    "source_url": ["source_url", "profile_url", "record_url"],
    "profile_url": ["profile_url", "public_profile_url"],
    "external_key": ["external_key", "external_id", "source_key"],
}


def read_rows(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))

    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not rows:
            return []

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        records = []
        for row in rows[1:]:
            if not any(value not in (None, "") for value in row):
                continue
            record = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = row[index] if index < len(row) else None
                record[header] = value
            records.append(record)
        return records

    if suffix == ".jsonl":
        rows = []
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
        return rows

    if suffix == ".json":
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if isinstance(payload, list):
            return payload
        if isinstance(payload, dict):
            for key in ["records", "items", "rows", "data"]:
                if isinstance(payload.get(key), list):
                    return payload[key]
        raise ValueError("JSON import expects a list or an object containing records/items/rows/data.")

    raise ValueError(f"Unsupported file type: {path.suffix}")


def load_field_map(path: Path | None):
    if not path:
        return {}
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise ValueError("Field map must be a JSON object.")
    return payload


def pick_value(row: dict, aliases):
    for alias in aliases:
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
    return None


def normalize_list(value):
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value)
    for separator in ["|", ";", ","]:
        if separator in text:
            return [part.strip() for part in text.split(separator) if part.strip()]
    return [text.strip()] if text.strip() else []


def build_canonical_row(row: dict, field_map: dict, dataset_name: str, row_number: int, default_state: str = "NY"):
    canonical = {}
    for field, aliases in DEFAULT_ALIASES.items():
        override = field_map.get(field)
        if isinstance(override, str):
            aliases = [override]
        elif isinstance(override, list) and override:
            aliases = override
        value = pick_value(row, aliases)
        if field in {"practice_areas", "certifications"}:
            canonical[field] = normalize_list(value)
        else:
            canonical[field] = value

    if not canonical.get("full_name"):
        full_name = " ".join(
            str(part).strip()
            for part in [canonical.get("first_name"), canonical.get("last_name")]
            if str(part or "").strip()
        )
        canonical["full_name"] = full_name or None

    canonical["state"] = canonical.get("state") or default_state
    canonical["external_key"] = canonical.get("external_key") or canonical.get("bar_number") or canonical.get("row_id") or f"{dataset_name}:{row_number}"
    canonical["row_id"] = canonical.get("row_id") or f"{dataset_name}:{row_number}"
    return canonical


def validate_canonical_row(canonical: dict):
    issues = []
    warnings = []

    if not canonical.get("full_name"):
        issues.append("missing full_name")
    if not canonical.get("bar_number"):
        warnings.append("missing bar_number")
    if not canonical.get("license_status"):
        warnings.append("missing license_status")
    if not canonical.get("city"):
        warnings.append("missing city")
    if not canonical.get("source_url") and not canonical.get("profile_url"):
        warnings.append("missing source_url/profile_url")

    return {"issues": issues, "warnings": warnings}
